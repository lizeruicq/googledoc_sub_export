from openpyxl import load_workbook

Lang_Dict = {
	"TW": True,
	"KR": True,
	"CN": True,
	"EN": True,
	"JP": True,
	"TH": True,
	"VN": True,
	"ID": True,
	"DE": True,
	"FR": True,
	"ES": True,
	"PT": True,
	"RU": True
}

count = 0
Lang = "all"
ExcelName = "0 .xlsx"
FileSuffix = ".srt"
TestFileSuffix = ".txt"
OutputPath = "Output/"


# def test():
# 	wb = load_workbook(ExcelName)
# 	sheetNames = wb.get_sheet_names()
# 	sheet = wb.get_sheet_by_name(sheetNames[0])
# 	procLangDict = {}
# 	for i in range(1, sheet.max_column + 1):
# 		potentialLang = sheet.cell(row=1, column=i).value
# 		if potentialLang in Lang_Dict:
# 			procLangDict[potentialLang] = i
# 	print(procLangDict)
#
# 	fileName = "test_"+ TestFileSuffix
# 	# f = open(fileName, 'w+')
# 	for i in range(2,sheet.max_row+1):
# 		startTime = sheet.cell(row=i, column=1).value
# 		endTime = sheet.cell(row=i, column=2).value
# 		if (not startTime is None) and (not endTime is None):
# 			srtStr = sheet.cell(row = i, column=3).value
# 			# f.write(str(i-1) + '\n')
# 			print("第" + str(i-1) + "行")
# 			# f.write(startTime + " --> " + endTime + '\n')
# 			print(startTime + " --> " + endTime )
# 			# f.write(srtStr + '\n\n')
# 			print(srtStr)
# 	print("中文字幕测试完成")

name = {
	"CN": ".zh_CN",
	"EN": ".en_US",
	"DE": ".de_DE",
	"FR": ".fr_FR",
	"TW": ".zh_TW",
	"KR": ".ko_KR",
	"TH": ".th_TH",
	"VN": ".vi_VN",
	"ID": ".id_ID",
	"JP": ".jp_JP",
	"PT": ".pt_PT",
	"RU": ".ru_RU",
	"ES": ".es_ES",
}

def CGSrtProc():
	number = input("输入对应序号来输出相应字幕:\n 1.CN  2.EN 3.DE 4.FR 5.TW 6.KR 7.TH 8.VN 9.ID 10.JP 11.PT 12.RU 13.ES 14.所有\n")
	map = \
		{
			"1": "CN",
			"2": "EN",
			"3": "DE",
			"4": "FR",
			"5": "TW",
			"6": "KR",
			"7": "TH",
			"8": "VN",
			"9": "ID",
			"10": "JP",
			"11": "PT",
			"12": "RU",
			"13": "ES",

			"14": "all",
		}

	if number not in map:
		print("你的输入有误，请重新输入!")
		print("-------------------------------------------------------------\n")
		CGSrtProc()

	Lang = map[number]
	createSrt(Lang)
	print("-------------------------------------------------------------\n")
	CGSrtProc()



def createSrt(Lang):
	wb = load_workbook(ExcelName)
	sheetNames = wb.sheetnames
	for i in range(len(sheetNames)):
		procSingleSheet(wb, sheetNames[i], Lang)

def procSingleSheet(wb, cgName, Lang):
	sheet = wb[cgName]
	procLangDict = {}
	for i in range(1, sheet.max_column + 1):
		potentialLang = sheet.cell(row = 1, column = i).value
		if (potentialLang in Lang_Dict):
			for j in range (2,sheet.max_row+1):
				srt = sheet.cell(row = j, column = i).value
				if srt != None:
					procLangDict[potentialLang] = i


	if Lang == "all":
		for key in procLangDict:
			procSingleSrt(sheet, cgName, key, procLangDict[key])
			print(cgName+ "_"+ key + "语言字幕输出完成")
	else:
		if Lang not in procLangDict:
			print("没有找到" + Lang+"的字幕" )
		else:
			procSingleSrt(sheet, cgName, Lang, procLangDict[Lang])
			print(cgName+ "_"+ Lang + "语言字幕输出完成")



def procSingleSrt(sheet, cgName, procLang, columnIndex):
	fileName = ""
	if Lang_Dict[procLang] == True:
		fileName = OutputPath + cgName + name[procLang] + FileSuffix
	else:
		fileName = OutputPath + cgName + FileSuffix

	f = open(fileName, 'w+', encoding='utf-8')
	for i in range(2, sheet.max_row+1):
		startTime = sheet.cell(row = i, column = 1).value
		endTime = sheet.cell(row = i, column = 2).value
		if (not startTime is None) and (not endTime is None):
			srtStr = sheet.cell(row = i, column = columnIndex).value
			if srtStr==None:
				srtStr = " "
			f.write(str(i-1) + '\n')
			f.write(startTime + " --> " + endTime + '\n')
			f.write(srtStr + '\n\n')
			# print(str(i-1))
			# print(startTime + " --> " + endTime)
			# print(srtStr)

	f.close()

def main():
	CGSrtProc()

if __name__ == '__main__':
	main()