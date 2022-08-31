import os
import sys

import gspread
import pythoncom
from openpyxl import load_workbook
from oauth2client.service_account import ServiceAccountCredentials
from gspread import utils
import cgsrtpy3
import win32com.client as win32

scopes = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

credentials = ServiceAccountCredentials.from_json_keyfile_name("credentials.json",
                                                               scopes)
file = gspread.authorize(credentials)
# sheet = file.open("游戏外")
# source = sheet.worksheet("字幕本地化")
# target = sheet.worksheet("带时间字幕")
# outputname = "默认名称"

# source_startcell = 'C2'
# source_endcell = 'K20'
# paraname = '参数.txt'

def findLocalSheet(paralist):
    # paralist = readpara.readpara(paraname)
    global source_startcell,source_endcell,sheet,outputname,localsh,local_maxrow,source,target
    print("读取输入参数...")
    xls = paralist[0]
    # xls = paralist[0] + '.xls'
    print("连接到谷歌文档...")
    try:
        sheet = file.open(paralist[1])
        source_startcell = paralist[4]#从C列开始
        source_endcell = paralist[5]
        outputname = paralist[6]

    # source = sheet.worksheet("字幕本地化")
    # target = sheet.worksheet("带时间字幕")
        source = sheet.worksheet(paralist[2])
        target = sheet.worksheet(paralist[3])
    except:
        input("连接到谷歌文档失败，请检查输入")
    # wblist = []
    # for root, dirs, files in os.walk(".", topdown=False):
    #     for name in files:
    #         if name.split('.')[-1] =='xls':
    #             wblist.append(name)

    # wbpath = os.path.abspath(wblist[0])
    print("读取本地文件...")

    wbpath = os.path.abspath(xls)

    arctime_wb_name = wbpath
    pythoncom.CoInitialize()
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        arctime_wb = excel.Workbooks.Open(arctime_wb_name)
    except:
        input("读取本地文件失败")
    arctime_wb.SaveAs(arctime_wb_name + "x", FileFormat=51)
    arctime_wb.Close()
    excel.Application.Quit()
    pythoncom.CoUninitialize()
    # wblist = []
    # for root, dirs, files in os.walk(".", topdown=False):
    #     for name in files:
    #         if name.split('.')[-1] == 'xlsx':
    #             wblist.append(name)
    #
    # arctime_wb = load_workbook(wblist[-1])
    arctime_wb = load_workbook(xls[:-4]+'.xlsx')


    localsh = arctime_wb[arctime_wb.sheetnames[0]]

    local_maxrow = localsh.max_row

    for i in range(2, local_maxrow+1):
        for j in range (2,4):
            a = localsh.cell(i,j).value
            b = a.replace('.',',')
            localsh.cell(i,j).value = b

    arctime_wb.save(xls[:-4]+'.xlsx')

    copyLocalSheet()

# source_row_count = len(source.col_values(1))
# source_col_count = len(source.row_values(1))


# print(source_row_count)
# print(source_col_count)

# target.batch_clear(["A2:K"+str(target.row_count)])
def copyLocalSheet():

    print("清扫目标表格的旧内容....")

    target.clear()

    print("拷贝时间轴到目标表格....")
    target.update("A1:B1",[['start','end']])
    locallist = []
    for i in range(2,local_maxrow+1):
        locallist.append([localsh.cell(i,2).value,localsh.cell(i,3).value])

    target.update("A2:B"+str(local_maxrow),locallist)
    copyGoogleSheet(source_startcell,source_endcell)

def copyGoogleSheet(startcell, endcell):
    print("寻找多语言字幕内容....")
    source_startcellpos = utils.a1_to_rowcol(startcell)
    source_endcellpos =  utils.a1_to_rowcol(endcell)

    hori_gap = source_endcellpos[1] - source_startcellpos[1]
    vert_gap = source_endcellpos[0] - source_startcellpos[0]

    srt_first_row = source.row_values(1)
    cur_firtst_row_num=len(target.row_values(1))


    trans_lang_list = srt_first_row[srt_first_row.index(source.acell(source_startcell[0]+'1').value):]
    srt_list = source.batch_get([startcell+':'+source_endcell])[0]

    print("计算表格填写位置....")
    target_startcell = utils.rowcol_to_a1(2,len(target.row_values(1))+1)
    target_endcell = utils.rowcol_to_a1(2+vert_gap, hori_gap+len(target.row_values(1))+1)

    target_firstlangcell = utils.rowcol_to_a1(1,cur_firtst_row_num+1)
    target_endlangcell = utils.rowcol_to_a1(1,cur_firtst_row_num+len(trans_lang_list))

    print("复制多语言字幕目标表格....")
    target.update(str(target_firstlangcell+":"+target_endlangcell),[trans_lang_list])
    target.update(str(target_startcell+":"+target_endcell),srt_list)

    convertTolLcal()


# for i in range(1,source_row_count):
#     if source.cell(i,1).value == "5.3PV":
#         for j in range(1,source_col_count+1):
#             if source.cell(1,j).value == "CN":
#                 print(source.cell(i,j).value)
                # CN_list.append([source.cell(i,j).value])
def convertTolLcal():
    print("将目标表格导出到本地...")
    content = target.get_all_values()

    swb = load_workbook("0 .xlsx")
    # ssh_name = source.acell('A'+str(source.acell(source_startcell).row)).value
    ssh_name = outputname
    ssh = swb.create_sheet(ssh_name,0)
    swb.remove(swb[swb.sheetnames[-1]])

    ssh_row = len(content)
    ssh_col = len(content[0])

# print(ssh_col,ssh_row)
    for i in range(1,ssh_row+1):
        for j in range(1,ssh_col+1):
            ssh.cell(i,j).value = content[i-1][j-1]

    swb.save("0 .xlsx")

    print("将表格导出为字幕文件...")
    cgsrtpy3.createSrt("all")
    input("操作完成！")
    exit()

# def main():
# 	findLocalSheet()

# if __name__ == '__main__':
# 	main()

